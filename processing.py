import tkinter as tk
from tkinter import filedialog
import pandas as pd
from openpyxl import load_workbook
from fuzzywuzzy import fuzz
from collections import defaultdict
from difflib import SequenceMatcher
import re

def has_multiple_sheets(file_path):
    workbook = load_workbook(file_path)
    return len(workbook.sheetnames) > 1

# Preprocess name function
def preprocess_name(name):
    if pd.isna(name):
        return ""
    name = str(name)
    words_to_omit = [
        'industry', 'industries', 'corp', 'corporation', 'inc', 'incorporated', 'foundation', 
        'company', 'co', 'limited', 'ltd', 'pvt', 'llc', 'llp', 'and', 'pvtltd', '&', 'm/s', 'ms'
    ]
    name = re.sub(r'\b(Mr|Ms|Ltd|LLP|Pvt|Private|Limited|LLC|LTD|Llp|ltd|lp|PLLP|Pllp|P.L.C.|ms|m/s|pvtltd)\b', '', name, flags=re.IGNORECASE)
    name = re.sub(r'\b(and|AND|And|&)\b', 'and', name, flags=re.IGNORECASE)
    name = re.sub(r'[.,]', '', name)
    for word in words_to_omit:
        name = re.sub(r'\b' + word + r'\b', '', name, flags=re.IGNORECASE)
    name = re.sub(r'\s+', '', name).lower()
    return name

# Compute similarity function
def compute_similarity(data1, data2, threshold=71):
    names1 = pd.Series(data1['CustName'].unique()).astype(str).apply(preprocess_name)
    names2 = pd.Series(data2['INSURED_CUSTOMER_NAME'].unique()).astype(str).apply(preprocess_name)
    results = {}
    index_dict_1 = defaultdict(list)
    index_dict_2 = defaultdict(list)
    
    for name1 in names1:
        for name2 in names2:
            similarity = fuzz.ratio(name1, name2)
            if similarity >= threshold:
                if similarity not in results:
                    results[similarity] = []
                results[similarity].append((name1, name2))
                index_dict_1[name1].extend(data1[data1['CustName'].apply(preprocess_name) == name1]['C.No.'].tolist())
                index_dict_2[name2].extend(data2[data2['INSURED_CUSTOMER_NAME'].apply(preprocess_name) == name2]['POL_NUM_TXT'].tolist())
    
    return results, index_dict_1, index_dict_2

def preprocess_text(text):
    if pd.isna(text):
        return ""
    return str(text).lower().strip()

def acronym(word):
    # Extract acronym from a phrase
    return ''.join([char for char in word if char.isupper()])

def clean_text(text):
    # Remove punctuation and make lowercase
    return re.sub(r'[^A-Za-z0-9\s]', '', text).lower()

def similarity(a, b):
    # Compute similarity between two strings
    return SequenceMatcher(None, a, b).ratio()

def find_similar_elements(list1, list2, threshold=0.75):
    results = []
    acronyms_list1 = [acronym(str(item)) for item in list1]
    acronyms_list2 = [acronym(str(item)) for item in list2]

    # Check for acronym matches
    for i, acr1 in enumerate(acronyms_list1):
        for j, acr2 in enumerate(acronyms_list2):
            if acr1 == acr2 and acr1:  # Only consider non-empty acronyms
                results.append((list1[i], list2[j]))

    # Check for similarity in remaining data
    for item1 in list1:
        for item2 in list2:
            cleaned_item1 = clean_text(str(item1))
            cleaned_item2 = clean_text(str(item2))
            if similarity(cleaned_item1, cleaned_item2) >= threshold:
                results.append((item1, item2))

    return results

def check_similarity_for_sorted_list(df1, df2, col1, col2, sorted_list, threshold=0.75):
    # Ensure we are working with copies to avoid SettingWithCopyWarning
    df1 = df1.copy()
    df2 = df2.copy()

    # Preprocess the policy names
    df1.loc[:, col1] = df1[col1].apply(preprocess_text)
    df2.loc[:, col2] = df2[col2].apply(preprocess_text)

    matched_sorted_list = []

    # Iterate through the sorted_list to check similarity for each pair
    for pair in sorted_list:
        index1 = list(pair.keys())[0]
        index2 = list(pair.values())[0]

        # Get the policy names corresponding to the indices
        policy_name_1 = df1.loc[df1['C.No.'] == index1, col1].values[0]
        policy_name_2 = df2.loc[df2['POL_NUM_TXT'] == index2, col2].values[0]

        # Combine the policy names for similarity checking
        combined_policies_1 = [policy_name_1]
        combined_policies_2 = [policy_name_2]

        # Check for similarity using the new logic
        similar_elements = find_similar_elements(combined_policies_1, combined_policies_2, threshold)

        # If any similar elements are found, keep the pair
        if similar_elements:
            matched_sorted_list.append(pair)

    return matched_sorted_list

# Premium Amount similarity check function
def is_within_2_percent(val1, val2):
    return abs(val1 - val2) <= 0.02 * val1
    
# Check Premium Amount similarity for updated sorted list
def check_premium_similarity(df1, df2, col1, col2, sorted_list):
    similar_pairs = []
    for pair in sorted_list:
        index1 = list(pair.keys())[0]
        index2 = list(pair.values())[0]
        
        premium1 = df1.loc[df1['C.No.'] == index1, col1].values[0]
        premium2 = df2.loc[df2['POL_NUM_TXT'] == index2, col2].values[0]
        
        if is_within_2_percent(premium1, premium2):
            similar_pairs.append(pair)
    
    return similar_pairs

# Preprocess name function
def preprocess_name(name):
    if pd.isna(name):
        return ""
    name = str(name)
    words_to_omit = [
        'industry', 'industries', 'corp', 'corporation', 'inc', 'incorporated', 'foundation', 
        'company', 'co', 'limited', 'ltd', 'pvt', 'llc', 'llp', 'and', 'pvtltd', '&', 'm/s', 'ms'
    ]
    name = re.sub(r'\b(Mr|Ms|Ltd|LLP|Pvt|Private|Limited|LLC|LTD|Llp|ltd|lp|PLLP|Pllp|P.L.C.|ms|m/s|pvtltd)\b', '', name, flags=re.IGNORECASE)
    name = re.sub(r'\b(and|AND|And|&)\b', 'and', name, flags=re.IGNORECASE)
    name = re.sub(r'[.,]', '', name)
    for word in words_to_omit:
        name = re.sub(r'\b' + word + r'\b', '', name, flags=re.IGNORECASE)
    name = re.sub(r'\s+', '', name).lower()
    return name

# Compute similarity function
def compute_similarity(data1, data2, threshold=71):
    names1 = pd.Series(data1['CustName'].unique()).astype(str).apply(preprocess_name)
    names2 = pd.Series(data2['INSURED_CUSTOMER_NAME'].unique()).astype(str).apply(preprocess_name)
    results = {}
    index_dict_1 = defaultdict(list)
    index_dict_2 = defaultdict(list)
    
    for name1 in names1:
        for name2 in names2:
            similarity = fuzz.ratio(name1, name2)
            if similarity >= threshold:
                if similarity not in results:
                    results[similarity] = []
                results[similarity].append((name1, name2))
                index_dict_1[name1].extend(data1[data1['CustName'].apply(preprocess_name) == name1]['C.No.'].tolist())
                index_dict_2[name2].extend(data2[data2['INSURED_CUSTOMER_NAME'].apply(preprocess_name) == name2]['POL_NUM_TXT'].tolist())
    
    return results, index_dict_1, index_dict_2

# Premium Amount similarity check function
def is_within_2_percent(val1, val2):
    return abs(val1 - val2) <= 0.02 * val1

# Check Premium Amount similarity for the sorted list
def check_premium_similarity(df1, df2, col1, col2, sorted_list):
    similar_pairs = []
    for pair in sorted_list:
        index1 = list(pair.keys())[0]
        index2 = list(pair.values())[0]
        
        premium1 = df1.loc[df1['C.No.'] == index1, col1].values[0]
        premium2 = df2.loc[df2['POL_NUM_TXT'] == index2, col2].values[0]
        
        if is_within_2_percent(premium1, premium2):
            similar_pairs.append(pair)
    
    return similar_pairs

def check_tenure_similarity(df1, df2, similarity_list):
    # Lists to store indices of matched policies
    Index1 = []
    Index2 = []

    # Iterate through each dictionary in the similarity list
    for pair in similarity_list:
        index1 = list(pair.keys())[0]
        index2 = list(pair.values())[0]

        # Retrieve the start and end dates from both dataframes
        start_date1 = df1.loc[df1['C.No.'] == index1, 'Policy_StartDate'].values[0]
        end_date1 = df1.loc[df1['C.No.'] == index1, 'Exp. Date'].values[0]
        start_date2 = df2.loc[df2['POL_NUM_TXT'] == index2, 'POLICY_START_DATE'].values[0]
        end_date2 = df2.loc[df2['POL_NUM_TXT'] == index2, 'POLICY_END_DATE'].values[0]

        # Check if both the start and end dates match
        if start_date1 == start_date2 and end_date1 == end_date2:
            Index1.append(index1)
            Index2.append(index2)
    
    return Index1, Index2

def process_excel(file1_path, file2_path):
    try:
        broker = pd.read_excel(file1_path)
    except Exception as e:
        print(f"Error reading {file1_path}: {e}")
        exit()

    try:
        if has_multiple_sheets(file2_path):
            company = pd.read_excel(file2_path, sheet_name='RAW STATEMENT')
        else:
            company = pd.read_excel(file2_path)
    except Exception as e:
        print(f"Error reading {file2_path}: {e}")
        exit()
        
    lombard = company.copy()
    lombard.insert(1, 'C.No.', '')
    lombard.insert(2, 'Inst No.', '')

    data1 = broker
    data2 = company

    saiba = data1.copy()

    c_num = [] #to store all the found control numbers

    #Checking Policy Number
    common_values_policy = set(data1['PolicyNo']).intersection(set(data2['POL_NUM_TXT']))
    for value in common_values_policy:
        c_no = data1.loc[data1['PolicyNo'] == value, 'C.No.'].values[0]
        c_num.append(int(c_no))
        lombard.loc[lombard['POL_NUM_TXT'] == value, 'C.No.'] = int(c_no)
        inst_no = data1.loc[data1['PolicyNo'] == value, 'Inst No.'].values[0]
        lombard.loc[lombard['POL_NUM_TXT'] == value, 'Inst No.'] = inst_no

    #Checking Endorsement Number
    common_values_endo = set(data1['EndoNo']).intersection(set(data2['POL_NUM_TXT']))
    for value in common_values_endo:
        c_no = data1.loc[data1['EndoNo'] == value, 'C.No.'].values
        c_num.append(int(c_no))
        lombard.loc[lombard['POL_NUM_TXT'] == value, 'C.No.'] = int(c_no)
        inst_no = data1.loc[data1['EndoNo'] == value, 'Inst No.'].values[0]
        lombard.loc[lombard['POL_NUM_TXT'] == value, 'Inst No.'] = inst_no
    
    data2 = lombard[~lombard['C.No.'].isin(c_num)]

    # Compute similarity with a threshold of 71%
    similarity_dict, index_dict_1, index_dict_2 = compute_similarity(data1, data2, threshold=71)

    # Sort the similarity_dict by keys in descending order
    sorted_similarity_dict = dict(sorted(similarity_dict.items(), key=lambda item: item[0], reverse=True))

    indexPairs = []
    for similarity, pairs in sorted_similarity_dict.items():
        for pair in pairs:
            for i in index_dict_1[pair[0]]:
                for j in index_dict_2[pair[1]]:
                    result_dict = {int(i): j}
                    if result_dict not in indexPairs:
                        indexPairs.append(result_dict)

    index_list_1 = [list(pair.keys())[0] for pair in indexPairs]
    index_list_2 = [list(pair.values())[0] for pair in indexPairs]
    filtered_data1 = data1[data1['C.No.'].isin(index_list_1)]
    filtered_data2 = data2[data2['POL_NUM_TXT'].isin(index_list_2)]

    # Compute updated sorted list based on policy name similarity
    updated_sorted_list = check_similarity_for_sorted_list(filtered_data1, filtered_data2, 'Policy Type', 'PRODUCT_NAME', indexPairs, threshold=0.75)

    # Compute premium similarity
    premium_similarity_list = check_premium_similarity(filtered_data1, filtered_data2, 'OD Premium', 'APPLICABLE_PREMIUM_AMOUNT', updated_sorted_list)

    for pairs in premium_similarity_list:
        lombard.loc[lombard['POL_NUM_TXT'] == list(pairs.values())[0], 'C.No.'] = list(pairs.keys())[0]
        c_num.append(list(pairs.keys())[0])
        inst_no = data1.loc[data1['C.No.'] == list(pairs.keys())[0], 'Inst No.'].values[0]
        lombard.loc[lombard['POL_NUM_TXT'] == list(pairs.values())[0], 'Inst No.'] = inst_no

    data2 = lombard[~lombard['C.No.'].isin(c_num)]
    data1 = saiba[~saiba['C.No.'].isin(c_num)]

    # Compute similarity with a threshold of 71%
    similarity_dict, index_dict_1, index_dict_2 = compute_similarity(data1, data2, threshold=71)

    # Sort the similarity_dict by keys in descending order
    sorted_similarity_dict = dict(sorted(similarity_dict.items(), key=lambda item: item[0], reverse=True))

    indexPairs = []
    for similarity, pairs in sorted_similarity_dict.items():
        for pair in pairs:
            for i in index_dict_1[pair[0]]:
                for j in index_dict_2[pair[1]]:
                    result_dict = {int(i): j}
                    if result_dict not in indexPairs:
                        indexPairs.append(result_dict)

    index_list_1 = [list(pair.keys())[0] for pair in indexPairs]
    index_list_2 = [list(pair.values())[0] for pair in indexPairs]
    filtered_data1 = data1[data1['C.No.'].isin(index_list_1)]
    filtered_data2 = data2[data2['POL_NUM_TXT'].isin(index_list_2)]

    # Compute premium similarity
    premium_similarity_list = check_premium_similarity(filtered_data1, filtered_data2, 'OD Premium', 'APPLICABLE_PREMIUM_AMOUNT', indexPairs)

    cont_no, pol_no = check_tenure_similarity(filtered_data1, filtered_data2, premium_similarity_list)

    for i in range(len(pol_no)):
        lombard.loc[lombard['POL_NUM_TXT'] == pol_no[i], 'C.No.'] = cont_no[i]
        c_num.append(cont_no[i])
        inst_no = data1.loc[data1['C.No.'] == cont_no[i], 'Inst No.'].values[0]
        lombard.loc[lombard['POL_NUM_TXT'] == pol_no[i], 'Inst No.'] = inst_no

    # Save and return output file path
    output_file_path = 'lombard.xlsx'
    lombard.to_excel(output_file_path, sheet_name='Lombard_Statement', index=False)
    return output_file_path